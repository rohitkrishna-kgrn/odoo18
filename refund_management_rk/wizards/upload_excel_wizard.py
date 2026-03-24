import base64
import io
from odoo import models, fields, _
from odoo.exceptions import UserError
from openpyxl import load_workbook

class UploadExcelWizard(models.TransientModel):
    _name = 'upload.excel.wizard'
    _description = 'Upload Excel for Reimbursement'

    excel_file = fields.Binary(string="Upload Excel", required=True)
    file_name = fields.Char(string="File Name")

    def action_process_excel(self):
        if not self.excel_file or not self.file_name:
            raise UserError(_("Please upload a valid Excel file."))

        try:
            file_data = base64.b64decode(self.excel_file)
            file_stream = io.BytesIO(file_data)

            # Important: reset the stream position
            file_stream.seek(0)

            wb = load_workbook(file_stream, data_only=True)
            ws = wb.active

            headers = {}
            data_lines = []
            first = True

            for row in ws.iter_rows(values_only=True):
                if first:
                    headers = {str(cell).strip().lower(): idx for idx, cell in enumerate(row)}
                    first = False
                    continue

                if all(cell is None for cell in row):
                    continue

                bill_date = row[headers.get('bill date')]
                purpose = row[headers.get('purpose')]
                bill_type = row[headers.get('type')]
                bill_submitted = row[headers.get('bill submitted')]
                bill_amount = row[headers.get('bill amount')]

                # Normalize boolean
                if isinstance(bill_submitted, str):
                    bill_submitted = bill_submitted.strip().lower() in ['true', 'yes', '1']
                else:
                    bill_submitted = bool(bill_submitted)

                data_lines.append({
                    'bill_date': bill_date,
                    'purpose': purpose,
                    'bill_type': bill_type.lower() if bill_type else False,
                    'bill_submitted': bill_submitted,
                    'bill_amount': float(bill_amount) if bill_amount else 0.0,
                })

        except Exception as e:
            raise UserError(_("Failed to read Excel file: %s") % str(e))

        # Pass parsed data to next wizard
        context = dict(self.env.context)
        context['parsed_lines'] = data_lines

        reimbursement_id = self.env.context.get('active_id')  # or another key if different
        if reimbursement_id:
            context['reimbursement_id'] = reimbursement_id
        return {
            'type': 'ir.actions.act_window',
            'name': 'Attach Bills',
            'res_model': 'attach.bills.wizard',
            'view_mode': 'form',
            'target': 'new',
            'context': context,
        }
